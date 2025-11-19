# app/services/export_service.py
"""
📄 Fichier: app/services/export_service.py
📝 Description: Service pour exports Excel et CSV
"""

from datetime import date
from typing import Optional
from io import BytesIO
import csv
from sqlalchemy.orm import Session
import pandas as pd

from app.models.cas import Cas
from app.models.alerte import Alerte
from app.models.intervention import Intervention
from app.models.maladie import Maladie
from app.models.district import District
from app.models.centre_sante import CentreSante


class ExportService:
    """Service pour les exports de données"""
    
    @staticmethod
    def export_cas_excel(
        db: Session,
        date_debut: Optional[date] = None,
        date_fin: Optional[date] = None,
        maladie_id: Optional[int] = None,
        district_id: Optional[int] = None
    ) -> BytesIO:
        """📊 Export des cas en Excel avec formatage"""
        
        # Query de base
        query = db.query(
            Cas.numero_cas,
            Cas.nom,
            Maladie.nom.label('maladie'),
            District.nom.label('district'),
            CentreSante.nom.label('centre_sante'),
            Cas.date_symptomes,
            Cas.date_declaration,
            Cas.age,
            Cas.sexe,
            Cas.statut,
            Cas.observations
        ).join(Maladie, Cas.maladie_id == Maladie.id)\
         .join(District, Cas.district_id == District.id)\
         .outerjoin(CentreSante, Cas.centre_sante_id == CentreSante.id)
        
        # Filtres
        if date_debut:
            query = query.filter(Cas.date_declaration >= date_debut)
        if date_fin:
            query = query.filter(Cas.date_declaration <= date_fin)
        if maladie_id:
            query = query.filter(Cas.maladie_id == maladie_id)
        if district_id:
            query = query.filter(Cas.district_id == district_id)
        
        # Récupération des données
        cas = query.all()
        
        # Conversion en DataFrame
        data = []
        for c in cas:
            data.append({
                'N° Cas': c.numero_cas,
                'Nom Patient': c.nom,
                'Maladie': c.maladie,
                'District': c.district,
                'Centre de Santé': c.centre_sante or 'N/A',
                'Date Symptômes': c.date_symptomes.strftime('%d/%m/%Y') if c.date_symptomes else 'N/A',
                'Date Déclaration': c.date_declaration.strftime('%d/%m/%Y') if c.date_declaration else 'N/A',
                'Âge': c.age or 'N/A',
                'Sexe': c.sexe or 'N/A',
                'Statut': c.statut.value if hasattr(c.statut, 'value') else str(c.statut),
                'Observations': c.observations or ''
            })
        
        df = pd.DataFrame(data)
        
        # Création du fichier Excel
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Cas', index=False)
            
            # Formatage de la feuille
            workbook = writer.book
            worksheet = writer.sheets['Cas']
            
            # Ajuster la largeur des colonnes
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Style de l'en-tête
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_cas_csv(
        db: Session,
        date_debut: Optional[date] = None,
        date_fin: Optional[date] = None,
        maladie_id: Optional[int] = None,
        district_id: Optional[int] = None
    ) -> BytesIO:
        """📊 Export des cas en CSV"""
        
        # Même query que pour Excel
        query = db.query(
            Cas.numero_cas,
            Cas.nom,
            Maladie.nom.label('maladie'),
            District.nom.label('district'),
            CentreSante.nom.label('centre_sante'),
            Cas.date_symptomes,
            Cas.date_declaration,
            Cas.age,
            Cas.sexe,
            Cas.statut,
            Cas.observations
        ).join(Maladie, Cas.maladie_id == Maladie.id)\
         .join(District, Cas.district_id == District.id)\
         .outerjoin(CentreSante, Cas.centre_sante_id == CentreSante.id)
        
        if date_debut:
            query = query.filter(Cas.date_declaration >= date_debut)
        if date_fin:
            query = query.filter(Cas.date_declaration <= date_fin)
        if maladie_id:
            query = query.filter(Cas.maladie_id == maladie_id)
        if district_id:
            query = query.filter(Cas.district_id == district_id)
        
        cas = query.all()
        
        # Création du CSV
        buffer = BytesIO()
        buffer.write('\ufeff'.encode('utf-8'))  # BOM pour Excel
        
        writer = csv.writer(buffer, delimiter=';')
        
        # En-tête
        writer.writerow([
            'N° Cas', 'Nom Patient', 'Maladie', 'District', 'Centre de Santé',
            'Date Symptômes', 'Date Déclaration', 'Âge', 'Sexe', 'Statut', 'Observations'
        ])
        
        # Données
        for c in cas:
            writer.writerow([
                c.numero_cas,
                c.nom,
                c.maladie,
                c.district,
                c.centre_sante or 'N/A',
                c.date_symptomes.strftime('%d/%m/%Y') if c.date_symptomes else 'N/A',
                c.date_declaration.strftime('%d/%m/%Y') if c.date_declaration else 'N/A',
                c.age or 'N/A',
                c.sexe or 'N/A',
                c.statut.value if hasattr(c.statut, 'value') else str(c.statut),
                c.observations or ''
            ])
        
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_alertes_excel(
        db: Session,
        date_debut: Optional[date] = None,
        date_fin: Optional[date] = None
    ) -> BytesIO:
        """🚨 Export des alertes en Excel"""
        
        query = db.query(
            Alerte.titre,
            Maladie.nom.label('maladie'),
            District.nom.label('district'),
            Alerte.niveau_gravite,
            Alerte.description,
            Alerte.date_detection,
            Alerte.statut,
            Alerte.nombre_cas
        ).join(Maladie, Alerte.maladie_id == Maladie.id)\
         .join(District, Alerte.district_id == District.id)
        
        if date_debut:
            query = query.filter(Alerte.date_detection >= date_debut)
        if date_fin:
            query = query.filter(Alerte.date_detection <= date_fin)
        
        alertes = query.all()
        
        data = []
        for a in alertes:
            data.append({
                'Titre': a.titre,
                'Maladie': a.maladie,
                'District': a.district,
                'Niveau': a.niveau_gravite.value if hasattr(a.niveau_gravite, 'value') else str(a.niveau_gravite),
                'Description': a.description,
                'Date Détection': a.date_detection.strftime('%d/%m/%Y'),
                'Statut': a.statut,
                'Nombre de Cas': a.nombre_cas or 0
            })
        
        df = pd.DataFrame(data)
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Alertes', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Alertes']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_interventions_excel(
        db: Session,
        date_debut: Optional[date] = None,
        date_fin: Optional[date] = None
    ) -> BytesIO:
        """🎯 Export des interventions en Excel"""
        
        query = db.query(
            Intervention.titre,
            Intervention.type,
            Maladie.nom.label('maladie'),
            District.nom.label('district'),
            Intervention.date_debut,
            Intervention.date_fin,
            Intervention.statut,
            Intervention.budget,
            Intervention.population_cible,
            Intervention.efficacite_score,
            Intervention.description
        ).join(Maladie, Intervention.maladie_id == Maladie.id)\
         .join(District, Intervention.district_id == District.id)
        
        if date_debut:
            query = query.filter(Intervention.date_debut >= date_debut)
        if date_fin:
            query = query.filter(Intervention.date_debut <= date_fin)
        
        interventions = query.all()
        
        data = []
        for i in interventions:
            data.append({
                'Titre': i.titre,
                'Type': i.type.value if hasattr(i.type, 'value') else str(i.type),
                'Maladie': i.maladie,
                'District': i.district,
                'Date Début': i.date_debut.strftime('%d/%m/%Y'),
                'Date Fin': i.date_fin.strftime('%d/%m/%Y') if i.date_fin else 'En cours',
                'Statut': i.statut,
                'Budget (Ar)': i.budget or 0,
                'Population Cible': i.population_cible or 0,
                'Score Efficacité': f"{i.efficacite_score}/5" if i.efficacite_score else 'N/A',
                'Description': i.description or ''
            })
        
        df = pd.DataFrame(data)
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Interventions', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Interventions']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        buffer.seek(0)
        return buffer


# Instance globale
export_service = ExportService()
