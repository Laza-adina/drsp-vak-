"""
📄 Fichier: app/api/v1/endpoints/export.py
📝 Description: Endpoints d'export de données
🎯 Usage: Export Excel et CSV des cas
"""

from datetime import date
from typing import Optional
from io import BytesIO, StringIO
import csv
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.api.deps import get_db, get_current_active_user
from app.crud.cas import cas as crud_cas
from app.models.user import User

router = APIRouter()


@router.get("/cas/excel")
def export_cas_excel(
    maladie_id: Optional[int] = Query(None),
    district_id: Optional[int] = Query(None),
    date_debut: Optional[date] = Query(None),
    date_fin: Optional[date] = Query(None),
    statut: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    📊 Exporte les cas au format Excel (.xlsx)
    
    Génère un fichier Excel formaté avec en-têtes stylisés (fond bleu DRSP),
    colonnes auto-ajustées, et toutes les données des cas filtrés.
    
    Limite : 10 000 cas par export pour éviter les timeouts.
    """
    # ✅ DÉJÀ CORRIGÉ : Suppression de .cas
    cas_list = crud_cas.get_by_filters(
        db=db,
        maladie_id=maladie_id,
        district_id=district_id,
        date_debut=date_debut,
        date_fin=date_fin,
        statut=statut,
        skip=0,
        limit=10000
    )
    
    # Créer le workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cas"
    
    # Style pour l'en-tête (bleu DRSP)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # En-têtes
    headers = [
        "N° Cas", "Maladie", "District", "Centre de Santé",
        "Date Symptômes", "Date Déclaration", "Âge", "Sexe",
        "Statut", "Latitude", "Longitude", "Observations"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Données
    for row_num, cas in enumerate(cas_list, 2):
        ws.cell(row=row_num, column=1, value=cas.numero_cas)
        ws.cell(row=row_num, column=2, value=cas.maladie.nom if cas.maladie else "")
        ws.cell(row=row_num, column=3, value=cas.district.nom if cas.district else "")
        ws.cell(row=row_num, column=4, value=cas.centre_sante.nom if cas.centre_sante else "")
        ws.cell(row=row_num, column=5, value=cas.date_symptomes.isoformat() if cas.date_symptomes else "")
        ws.cell(row=row_num, column=6, value=cas.date_declaration.isoformat() if cas.date_declaration else "")
        ws.cell(row=row_num, column=7, value=cas.age)
        ws.cell(row=row_num, column=8, value=cas.sexe if cas.sexe else "")
        ws.cell(row=row_num, column=9, value=cas.statut if cas.statut else "")
        ws.cell(row=row_num, column=10, value=cas.latitude)
        ws.cell(row=row_num, column=11, value=cas.longitude)
        ws.cell(row=row_num, column=12, value=cas.observations)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"cas_export_{date.today().isoformat()}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/cas/csv")
def export_cas_csv(
    maladie_id: Optional[int] = Query(None),
    district_id: Optional[int] = Query(None),
    date_debut: Optional[date] = Query(None),
    date_fin: Optional[date] = Query(None),
    statut: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    📄 Exporte les cas au format CSV
    
    Génère un fichier CSV simple compatible avec Excel, LibreOffice,
    et tous les logiciels d'analyse de données.
    
    Format : UTF-8, séparateur virgule, encodage standard.
    Limite : 10 000 cas par export.
    """
    # ✅ DÉJÀ CORRIGÉ : Suppression de .cas
    cas_list = crud_cas.get_by_filters(
        db=db,
        maladie_id=maladie_id,
        district_id=district_id,
        date_debut=date_debut,
        date_fin=date_fin,
        statut=statut,
        skip=0,
        limit=10000
    )
    
    # Créer le CSV
    output = StringIO()
    writer = csv.writer(output)
    
    # En-têtes
    writer.writerow([
        "numero_cas", "maladie", "district", "centre_sante",
        "date_symptomes", "date_declaration", "age", "sexe",
        "statut", "latitude", "longitude", "observations"
    ])
    
    # Données
    for cas in cas_list:
        writer.writerow([
            cas.numero_cas,
            cas.maladie.nom if cas.maladie else "",
            cas.district.nom if cas.district else "",
            cas.centre_sante.nom if cas.centre_sante else "",
            cas.date_symptomes.isoformat() if cas.date_symptomes else "",
            cas.date_declaration.isoformat() if cas.date_declaration else "",
            cas.age,
            cas.sexe if cas.sexe else "",
            cas.statut if cas.statut else "",
            cas.latitude,
            cas.longitude,
            cas.observations
        ])
    
    output.seek(0)
    filename = f"cas_export_{date.today().isoformat()}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
